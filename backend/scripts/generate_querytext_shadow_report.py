from __future__ import annotations

import json
import os
import sys

from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from fastapi.testclient import TestClient
from openpyxl.styles import Alignment, Font, PatternFill
from qdrant_client import QdrantClient

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend" / "src"))

from inrules_data_agent.app import create_app  # noqa: E402

SOURCE = ROOT / os.environ.get(
    "SHADOW_REPORT_SOURCE",
    "Decompose_query_review_20260716_PIPELINE_VALIDATED_V9_WITH_ORIGINAL_QUERY.xlsx",
)
OUTPUT = ROOT / os.environ.get(
    "SHADOW_REPORT_OUTPUT",
    "Decompose_query_review_20260716_PIPELINE_VALIDATED_V9_QUERYTEXT_SHADOW.xlsx",
)
PAYLOAD_SHEET = os.environ.get("SHADOW_REPORT_PAYLOAD_SHEET", "Payloads V9")
REPORT_SHEET = "QueryText Shadow Report"
RESPONSE_SHEET = "QueryText Backend Responses"
CANDIDATE_SHEET = "QueryText Candidate Comparison"


def require_qdrant() -> None:
    if os.environ.get("QDRANT_ENABLED", "false").lower() not in {"1", "true", "yes"}:
        raise RuntimeError("QDRANT_ENABLED must be true for this local report")
    url = os.environ.get("QDRANT_URL", "http://localhost:6333")
    collection = os.environ.get("QDRANT_COLLECTION", "inrule_schema")
    info = QdrantClient(url=url).get_collection(collection)
    if not info.points_count:
        raise RuntimeError(f"Qdrant collection {collection} is empty")
    print(f"Qdrant ready: {collection} ({info.points_count} points)", flush=True)


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    match_fill = PatternFill("solid", fgColor="C6EFCE")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in range(2, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            sheet.cell(row, column).alignment = Alignment(vertical="top", wrap_text=True)
        if sheet.cell(row, 5).value == "Y":
            for column in range(1, sheet.max_column + 1):
                sheet.cell(row, column).fill = match_fill
        sheet.row_dimensions[row].height = 180
    widths = [12, 11, 48, 75, 14, 14, 28, 55, 35, 55]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions


def main() -> None:
    load_dotenv(ROOT / ".env")
    load_dotenv(ROOT / "backend" / ".env")
    require_qdrant()
    os.environ["DATAQUERY_SHADOW_ENABLED"] = "true"

    workbook = openpyxl.load_workbook(SOURCE)
    payload_sheet = workbook[PAYLOAD_SHEET]
    for name in (REPORT_SHEET, CANDIDATE_SHEET, RESPONSE_SHEET):
        if name in workbook.sheetnames:
            del workbook[name]

    report = workbook.create_sheet(REPORT_SHEET)
    report.append(
        [
            "Rule ID",
            "Step #",
            "Business Meaning",
            "Generated Query",
            "QueryText Match",
            "DataQueryId",
            "DataQuery Name",
            "Stored QueryText",
            "Matched Tables",
            "Normalized Predicates",
        ]
    )
    candidates_sheet = workbook.create_sheet(CANDIDATE_SHEET)
    candidates_sheet.append([
        "Rule ID", "Step #", "Business Meaning", "Generated Query", "Candidate DataQueryId",
        "Candidate Name", "Stored QueryText", "Matched Tables", "Candidate Filter Columns",
        "Common Filter Columns", "Missing Filter Columns", "Extra Filter Columns", "Strict Filter Set",
    ])
    responses = workbook.create_sheet(RESPONSE_SHEET)
    responses.append(["Rule ID", "Backend Response JSON"])

    client = TestClient(create_app())
    total_steps = total_matches = 0
    for row in range(2, payload_sheet.max_row + 1):
        rule_id = str(payload_sheet.cell(row, 1).value or "").strip()
        raw_payload = payload_sheet.cell(row, 2).value
        if not rule_id or not raw_payload:
            continue
        payload = json.loads(str(raw_payload))
        response = client.post("/generate_queries", json=payload)
        response.raise_for_status()
        body = response.json()
        responses.append([rule_id, json.dumps(body, indent=2, ensure_ascii=False)])
        for item in body.get("queries", []):
            total_steps += 1
            generated = "\n\n".join(item.get("queries") or [])
            matches = item.get("querytext_shadow_matches") or []
            for candidate in item.get("querytext_comparison_candidates") or []:
                candidates_sheet.append([
                    rule_id, item.get("step_number"), item.get("business_meaning"), generated,
                    candidate.get("data_query_id"), candidate.get("name"), candidate.get("query_text"),
                    "\n".join(candidate.get("tables") or []),
                    "\n".join(candidate.get("filter_columns") or []),
                    "\n".join(candidate.get("common_filter_columns") or []),
                    "\n".join(candidate.get("missing_filter_columns") or []),
                    "\n".join(candidate.get("extra_filter_columns") or []),
                    "Y" if candidate.get("strict_match") else "N",
                ])
            if not matches:
                report.append(
                    [rule_id, item.get("step_number"), item.get("business_meaning"), generated, "N", None, None, None, None, None]
                )
                continue
            for match in matches:
                total_matches += 1
                basis = match.get("match_basis") or {}
                report.append(
                    [
                        rule_id,
                        item.get("step_number"),
                        item.get("business_meaning"),
                        generated,
                        "Y",
                        match.get("data_query_id"),
                        match.get("name"),
                        match.get("query_text"),
                        "\n".join(basis.get("tables") or []),
                        basis.get("normalized_predicates"),
                    ]
                )
        workbook.save(OUTPUT)
        print(f"rule={rule_id} complete", flush=True)

    style_sheet(report)
    for cell in candidates_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in range(2, candidates_sheet.max_row + 1):
        for column in range(1, candidates_sheet.max_column + 1):
            candidates_sheet.cell(row, column).alignment = Alignment(vertical="top", wrap_text=True)
        candidates_sheet.row_dimensions[row].height = 150
    for index, width in enumerate([12, 10, 42, 60, 15, 26, 55, 28, 28, 25, 25, 25, 15], 1):
        candidates_sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    candidates_sheet.freeze_panes = "C2"
    candidates_sheet.auto_filter.ref = candidates_sheet.dimensions
    responses.column_dimensions["A"].width = 14
    responses.column_dimensions["B"].width = 120
    for cell in responses[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in range(2, responses.max_row + 1):
        responses.cell(row, 2).alignment = Alignment(vertical="top", wrap_text=True)
        responses.row_dimensions[row].height = 180
    workbook.save(OUTPUT)
    print(json.dumps({"output": str(OUTPUT), "steps": total_steps, "matches": total_matches}, indent=2), flush=True)


if __name__ == "__main__":
    main()
