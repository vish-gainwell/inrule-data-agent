from __future__ import annotations

import json
import os
from collections import Counter, defaultdict
from pathlib import Path
from typing import cast

import openpyxl
import pyodbc
from dotenv import load_dotenv
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from inrules_data_agent.retrieval.select_inventory import SelectEvidence, analyze_select, load_frontier_tables

ROOT = Path(__file__).resolve().parents[2]
BACKEND = ROOT / "backend"
OUTPUT = ROOT / "Select_QueryText_Global_Inventory.xlsx"
FRONTIER_DIR = BACKEND / "src" / "inrules_data_agent" / "in_memory_schema"


def connection_string() -> str:
    load_dotenv(ROOT / ".env")
    hostname = os.environ.get("DB_HOSTNAME") or os.environ.get("hostname")
    username = os.environ.get("DB_USERNAME") or os.environ.get("db_username")
    password = os.environ.get("DB_PASSWORD") or os.environ.get("db_password")
    port = os.environ.get("DB_PORT") or os.environ.get("port") or "1433"
    trust = os.environ.get("DB_TRUST_SERVER_CERTIFICATE", "yes")
    if not hostname or not username or not password:
        raise RuntimeError("ClaimEngine database credentials are not configured")
    return (
        "DRIVER={ODBC Driver 18 for SQL Server};"
        f"SERVER={hostname},{port};DATABASE=master;UID={username};PWD={password};"
        f"Encrypt=yes;TrustServerCertificate={trust}"
    )


def load_rows() -> list[dict[str, object]]:
    query = """
        SELECT
            dp.DataPackageId,
            dp.DataPackageName,
            dp.DataPackageVersion,
            dp.EditId,
            dp.Active AS PackageActive,
            dpdq.Id AS AssignmentId,
            dpdq.Active AS AssignmentActive,
            dpdq.Priority,
            dpdq.QueryParams,
            dpdq.ReturnVals,
            dpdq.IsDataSet,
            dq.DataQueryId,
            dq.Name AS DataQueryName,
            dq.DbId,
            dq.QueryType,
            dq.QueryText
        FROM ClaimEngine.dbo.DataPackage dp
        JOIN ClaimEngine.dbo.DataPackageDataQuery dpdq
          ON dpdq.DataPackage_Id = dp.DataPackageId
        JOIN ClaimEngine.dbo.DataQuery dq
          ON dq.DataQueryId = dpdq.DataQuery_Id
        WHERE LOWER(LTRIM(RTRIM(dq.QueryType))) = 'select'
          AND dq.QueryText IS NOT NULL
        ORDER BY dp.EditId, dp.DataPackageVersion, dpdq.Priority, dpdq.Id
    """
    with pyodbc.connect(connection_string(), timeout=30) as connection:
        cursor = connection.cursor()
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def parse_json(value: object, fallback: object) -> object:
    try:
        return json.loads(str(value or ""))
    except (json.JSONDecodeError, TypeError):
        return fallback


def enrich(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    frontier = load_frontier_tables(FRONTIER_DIR)
    enriched: list[dict[str, object]] = []
    for source_row in rows:
        row = dict(source_row)
        evidence = analyze_select(str(row["QueryText"]), frontier)
        params = parse_json(row.get("QueryParams"), {})
        returns = parse_json(row.get("ReturnVals"), [])
        mapped_names = {
            str(name).lstrip(":").lower()
            for name in params
        } if isinstance(params, dict) else set()
        unresolved = tuple(
            name for name in evidence.placeholders if name.lower() not in mapped_names
        )
        row.update(
            {
                "CurrentAssignment": bool(row["PackageActive"] and row["AssignmentActive"]),
                "ParseStatus": evidence.parse_status,
                "QueryShape": evidence.query_shape,
                "SourceClassification": evidence.source,
                "Tables": "\n".join(dict.fromkeys(table.raw_name for table in evidence.tables)),
                "TableEvidence": "\n".join(
                    f"{table.raw_name}: {table.source} ({table.basis})" for table in evidence.tables
                ),
                "Joins": "\n".join(evidence.joins),
                "FilterColumns": "\n".join(evidence.filter_columns),
                "Placeholders": "\n".join(evidence.placeholders),
                "UnmappedPlaceholders": "\n".join(unresolved),
                "HasDynamicOutput": evidence.has_dynamic_output,
                "HasEffectiveDateFilter": evidence.has_effective_date_filter,
                "HasOrderBy": evidence.has_order_by,
                "StatementCount": evidence.statement_count,
                "ParseError": evidence.parse_error,
                "ParameterMappings": "\n".join(
                    f"{key} = {value}" for key, value in params.items()
                ) if isinstance(params, dict) else str(row.get("QueryParams") or ""),
                "ReturnedVariables": "\n".join(str(value) for value in returns)
                if isinstance(returns, list) else str(row.get("ReturnVals") or ""),
                "_Evidence": evidence,
            }
        )
        enriched.append(row)
    return enriched


def workbook_value(value: object) -> object:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def add_sheet(workbook: openpyxl.Workbook, name: str, columns: list[str], rows: list[dict[str, object]]) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(columns)
    for row in rows:
        sheet.append([workbook_value(row.get(column)) for column in columns])


def build_workbook(rows: list[dict[str, object]]) -> dict[str, int]:
    current = [row for row in rows if row["CurrentAssignment"]]
    historical = [row for row in rows if not row["CurrentAssignment"]]
    summary: dict[str, int] = {
        "all_select_assignments": len(rows),
        "current_select_assignments": len(current),
        "historical_select_assignments": len(historical),
        "current_edits": len({row["EditId"] for row in current}),
        "current_unique_dataqueries": len({row["DataQueryId"] for row in current}),
        "current_inmemory_assignments": sum(row["SourceClassification"] == "INMEMORY" for row in current),
        "current_physical_assignments": sum(row["SourceClassification"] == "PHYSICAL" for row in current),
        "current_mixed_assignments": sum(row["SourceClassification"] == "MIXED" for row in current),
        "current_expression_selects": sum(row["QueryShape"] == "EXPRESSION_SELECT" for row in current),
        "current_multi_table_selects": sum(row["QueryShape"] == "MULTI_TABLE_SELECT" for row in current),
        "current_parse_or_shape_issues": sum(row["ParseStatus"] != "PARSED" for row in current),
    }

    workbook = openpyxl.Workbook()
    overview = cast(Worksheet, workbook.active)
    overview.title = "Summary"
    overview.append(["Metric", "Value"])
    for key, value in summary.items():
        overview.append([key, value])

    detail_columns = [
        "EditId", "DataPackageId", "DataPackageVersion", "PackageActive", "AssignmentId",
        "AssignmentActive", "Priority", "DataQueryId", "DataQueryName", "DbId", "QueryShape",
        "SourceClassification", "Tables", "TableEvidence", "Joins", "FilterColumns",
        "HasDynamicOutput", "HasEffectiveDateFilter", "HasOrderBy", "Placeholders",
        "UnmappedPlaceholders", "ParameterMappings", "ReturnedVariables", "IsDataSet",
        "ParseStatus", "ParseError", "QueryText",
    ]
    add_sheet(workbook, "Current SELECT Assignments", detail_columns, current)
    add_sheet(workbook, "Historical SELECT Assignments", detail_columns, historical)

    edit_groups: dict[object, list[dict[str, object]]] = defaultdict(list)
    for row in current:
        edit_groups[row["EditId"]].append(row)
    edit_rows = []
    for edit_id, items in sorted(edit_groups.items(), key=lambda pair: str(pair[0])):
        edit_rows.append(
            {
                "EditId": edit_id,
                "ActiveSelectAssignments": len(items),
                "UniqueDataQueries": len({item["DataQueryId"] for item in items}),
                "InMemory": sum(item["SourceClassification"] == "INMEMORY" for item in items),
                "Physical": sum(item["SourceClassification"] == "PHYSICAL" for item in items),
                "Mixed": sum(item["SourceClassification"] == "MIXED" for item in items),
                "Expressions": sum(item["QueryShape"] == "EXPRESSION_SELECT" for item in items),
                "MultiTable": sum(item["QueryShape"] == "MULTI_TABLE_SELECT" for item in items),
                "Tables": "\n".join(sorted({table.raw_name for item in items for table in cast(SelectEvidence, item["_Evidence"]).tables})),
                "DataQueryIds": "\n".join(
                    sorted(
                        {str(item["DataQueryId"]) for item in items},
                        key=lambda value: int(value),
                    )
                ),
            }
        )
    add_sheet(
        workbook,
        "Current Edit Summary",
        ["EditId", "ActiveSelectAssignments", "UniqueDataQueries", "InMemory", "Physical", "Mixed", "Expressions", "MultiTable", "Tables", "DataQueryIds"],
        edit_rows,
    )

    table_counts: Counter[tuple[str, str, str]] = Counter()
    for row in current:
        evidence = cast(SelectEvidence, row["_Evidence"])
        for table in {item.normalized_name: item for item in evidence.tables}.values():
            table_counts[(table.raw_name, table.source, table.basis)] += 1
    table_rows = [
        {"Table": key[0], "Source": key[1], "Basis": key[2], "AssignmentCount": count}
        for key, count in table_counts.most_common()
    ]
    add_sheet(workbook, "Current Table Frequency", ["Table", "Source", "Basis", "AssignmentCount"], table_rows)

    query_groups: dict[object, list[dict[str, object]]] = defaultdict(list)
    for row in current:
        query_groups[row["DataQueryId"]].append(row)
    reuse_rows = []
    for data_query_id, items in query_groups.items():
        reuse_rows.append(
            {
                "DataQueryId": data_query_id,
                "DataQueryName": items[0]["DataQueryName"],
                "AssignmentCount": len(items),
                "EditCount": len({item["EditId"] for item in items}),
                "Edits": "\n".join(sorted({str(item["EditId"]) for item in items})),
                "QueryShape": items[0]["QueryShape"],
                "SourceClassification": items[0]["SourceClassification"],
                "QueryText": items[0]["QueryText"],
            }
        )
    reuse_rows.sort(key=lambda row: (-int(row["AssignmentCount"]), int(row["DataQueryId"])))
    add_sheet(
        workbook,
        "Current DataQuery Reuse",
        ["DataQueryId", "DataQueryName", "AssignmentCount", "EditCount", "Edits", "QueryShape", "SourceClassification", "QueryText"],
        reuse_rows,
    )

    join_counts: Counter[str] = Counter()
    for row in current:
        join_counts.update(cast(SelectEvidence, row["_Evidence"]).joins)
    join_rows = [{"JoinPattern": join, "AssignmentCount": count} for join, count in join_counts.most_common()]
    add_sheet(workbook, "Current Join Patterns", ["JoinPattern", "AssignmentCount"], join_rows)

    issue_rows = [
        row for row in rows
        if row["ParseStatus"] != "PARSED" or row["UnmappedPlaceholders"]
    ]
    add_sheet(
        workbook,
        "Review Candidates",
        ["EditId", "CurrentAssignment", "AssignmentId", "DataQueryId", "DataQueryName", "ParseStatus", "QueryShape", "UnmappedPlaceholders", "QueryParams", "QueryText", "ParseError"],
        issue_rows,
    )

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in range(1, sheet.max_column + 1):
            values = [len(str(sheet.cell(row, column).value or "")) for row in range(1, min(sheet.max_row, 200) + 1)]
            width = min(70, max(values, default=10) + 2)
            sheet.column_dimensions[get_column_letter(column)].width = max(12, width)

    workbook.save(OUTPUT)
    return summary


def main() -> None:
    rows = enrich(load_rows())
    summary = build_workbook(rows)
    print(json.dumps({"output": str(OUTPUT), **summary}, indent=2))


if __name__ == "__main__":
    main()
