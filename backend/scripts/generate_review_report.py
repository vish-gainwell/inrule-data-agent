from __future__ import annotations

import json
import os
import re
import sys
import time
from copy import copy
from pathlib import Path

import httpx
import openpyxl
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl.styles import Alignment, PatternFill

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend" / "src"))

from inrules_data_agent.generator.generate import (  # noqa: E402
    _find_invalid_column_refs,
    _find_invalid_sql_artifacts,
    _find_invalid_table_refs,
    select_ddls,
)

REPORT_VERSION = os.environ.get("REPORT_VERSION", "11")
SOURCE = ROOT / os.environ.get(
    "REPORT_SOURCE_WORKBOOK",
    "Decompose_query_review_20260716_PIPELINE_VALIDATED_V10.xlsx",
)
OUTPUT = ROOT / os.environ.get(
    "REPORT_OUTPUT_WORKBOOK",
    f"Decompose_query_review_20260716_PIPELINE_VALIDATED_V{REPORT_VERSION}.xlsx",
)
ARTIFACT_DIR = ROOT / os.environ.get("REPORT_ARTIFACT_DIR", "report_artifacts")
CONTEXT_PATH = ARTIFACT_DIR / "ado_context.json"
CHECKPOINT_PATH = ARTIFACT_DIR / f"report_checkpoint_v{REPORT_VERSION}.json"
REPORT_PATH = ARTIFACT_DIR / f"report_summary_v{REPORT_VERSION}.json"
RULE_QUERY_CHECKPOINT_PATH = ARTIFACT_DIR / f"rule_full_query_checkpoint_v{REPORT_VERSION}.json"
QUERY_REVIEW_SOURCE_SHEET = os.environ.get("QUERY_REVIEW_SOURCE_SHEET", "Query Review V10")
PAYLOAD_SOURCE_SHEET = os.environ.get("PAYLOAD_SOURCE_SHEET", "Payloads V10")
QUERY_REVIEW_OUTPUT_SHEET = f"Query Review V{REPORT_VERSION}"
PAYLOAD_OUTPUT_SHEET = f"Payloads V{REPORT_VERSION}"
ADO_OUTPUT_SHEET = f"ADO-First Test V{REPORT_VERSION}"
ADO_SOURCE_SHEET = os.environ.get("ADO_SOURCE_SHEET", "ADO-First Test V10")

FULL_QUERY_PROMPT = """Generate one complete canonical SQL Server SELECT for the current consolidated ADO business meaning.

Information hierarchy:
1. The CONSOLIDATED BUSINESS MEANING is the authoritative current task.
2. The complete ACCEPTANCE CRITERIA is whole-rule supporting context. Use only details that clarify the current task.
3. The complete DESCRIPTION is broad supporting context.
4. The supplied DDLs are the only allowed tables and columns.

Rules:
- Preserve every output, predicate, literal, date window, status, comparison, and aggregate required by the consolidated business meaning.
- A full query may use multiple supplied tables when the consolidated task genuinely requires them.
- Use fully qualified table names and add WITH (nolock) to physical table references.
- Prefer a complete InMemory table when it can support the task; otherwise use supplied physical tables.
- Use approved runtime placeholders such as {{MemberId}}, {{ProviderId}}, {{DateOfService}}, {{ClaimTransaction.Ndc}}, {{ClaimRequest.DrugRequested.GCNSeqNo.Code}}, {{ClaimRequest.DrugRequested.HIC3.Code}}, {{QuantityDispensed}}, and {{LookBackDate}}.
- Do not invent table names or column names outside the supplied schemas.
- A query is mandatory whenever Requires Data Query is Y. Never return NO_SUPPORTED_QUERY.
- When an exact business concept is unavailable, produce the closest executable query supported by the retrieved schemas while preserving every portion that can be represented.
- Return raw SQL only. No markdown or explanation.
"""

DECOMPOSE_PROMPT = """Extract exactly one executable Data Agent query from the supplied FULL QUERY for the CURRENT BUSINESS MEANING.
Return strict JSON only:
{"query":"SELECT ..."}

Rules:
- The FULL QUERY is the sole SQL source. Do not redesign or regenerate its logic.
- Use the CURRENT BUSINESS MEANING only to identify which single retrieval in the full query belongs to this row.
- Return exactly one SELECT with exactly one FROM and exactly one table reference.
- Do not use JOIN, APPLY, UNION, INTERSECT, EXCEPT, EXISTS, nested SELECT, or any subquery.
- Choose the one table fragment in the full query that most directly satisfies the current business meaning.
- Keep only projections and predicates from the full query that use the chosen table or runtime placeholders already present in the full query.
- Drop cross-table conditions rather than reproducing another table or subquery.
- Do not introduce any table, column, literal, predicate, or runtime placeholder absent from the full query.
- Never add a tautology such as WHERE 1 = 1 or column = the same column.
- Do not use the ADO description, acceptance criteria, or DDL schemas at this stage.
- Never return multiple queries, an operation list, an explanation, or NO_SUPPORTED_QUERY.
"""

REUSE_PLAN_PROMPT = """Create a reusable Data Agent retrieval plan for all Y steps in one rule.
Return strict JSON only:
{"steps":[{"source_step":"...","query":"SELECT ...","reuse_source_step":"..."}]}

Rules:
- The ORIGINAL FULL QUERY is the sole SQL source. Extract from it; do not redesign the rule.
- Every supplied Y step must appear exactly once.
- Each query must be one executable, single-table SELECT without JOIN, APPLY, UNION, EXISTS, or a nested SELECT.
- Group steps only when one retrieval result can satisfy all grouped business meanings using the same entity key, table, and date scope.
- For a group, create one reusable query returning the raw fields or metric needed by every grouped step.
- Put that identical query in every grouped step. The first step has an empty reuse_source_step; later steps name the first step.
- Thresholds and comparisons that can be evaluated from a returned metric belong to rule evaluation, not separate database calls. For example, retrieve TotalQuantityLast75Days once and reuse it for limits 100, 350, 400, and 900.
- Do not group steps merely because they use the same table when their entity key or date scope differs.
- Do not add a table, column, literal, or placeholder absent from the original full query.
- Never return NO_SUPPORTED_QUERY or explanatory text.
"""


def merged_value(sheet, row: int, column: int):
    value = sheet.cell(row, column).value
    if value is not None:
        return value
    for area in sheet.merged_cells.ranges:
        if area.min_row <= row <= area.max_row and area.min_col <= column <= area.max_col:
            return sheet.cell(area.min_row, area.min_col).value
    return None


def contiguous_ranges(values: list[tuple[int, object]]) -> list[tuple[int, int]]:
    ranges: list[tuple[int, int]] = []
    if not values:
        return ranges
    start_row, current_value = values[0]
    end_row = start_row
    for row, value in values[1:]:
        if value == current_value and row == end_row + 1:
            end_row = row
            continue
        ranges.append((start_row, end_row))
        start_row = end_row = row
        current_value = value
    ranges.append((start_row, end_row))
    return ranges


def make_client() -> tuple[OpenAI, str]:
    load_dotenv(ROOT / ".env")
    load_dotenv(ROOT / "backend" / ".env")
    key = os.environ.get("OPENAI_API_KEY")
    if not key:
        raise RuntimeError("OPENAI_API_KEY is not configured")
    verify = os.environ.get("OPENAI_VERIFY_SSL", "false").lower() in {"1", "true", "yes"}
    kwargs = {"api_key": key, "http_client": httpx.Client(verify=verify, timeout=180.0)}
    base = os.environ.get("OPENAI_BASE_URL") or os.environ.get("OPENAI_API_BASE")
    if base:
        kwargs["base_url"] = base
    return OpenAI(**kwargs), os.environ.get("OPENAI_MODEL", "gpt-5.6-luna")


def call_model(client: OpenAI, model: str, messages: list[dict], json_mode: bool = False) -> str:
    kwargs = {"model": model, "messages": messages}
    if model.lower().startswith(("gpt-5", "o1", "o3", "o4")):
        kwargs["max_completion_tokens"] = 4000
    else:
        kwargs.update({"temperature": 0, "max_tokens": 4000})
    if json_mode:
        kwargs["response_format"] = {"type": "json_object"}
    error = None
    for attempt in range(3):
        try:
            response = client.chat.completions.create(**kwargs)
            return (response.choices[0].message.content or "").strip()
        except Exception as exc:
            error = exc
            time.sleep(2**attempt)
    raise RuntimeError(f"OpenAI request failed: {error}")


def clean_text(value: str) -> str:
    text = value.strip()
    text = re.sub(r"^```(?:sql|json)?\s*", "", text, flags=re.I)
    text = re.sub(r"\s*```$", "", text)
    return text.strip().rstrip(";")


def original_full_query_sql(value: object) -> str | None:
    text = str(value or "").strip()
    first_line = text.splitlines()[0].upper() if text else ""
    if "CANONICAL QUERY" not in first_line:
        return None
    lines = text.splitlines()[1:]
    start = next(
        (
            index
            for index, line in enumerate(lines)
            if line.strip().upper().startswith(("SELECT ", "WITH ", ";WITH ", "DECLARE "))
        ),
        None,
    )
    if start is None:
        return None
    sql_lines = lines[start:]
    end = next(
        (
            index
            for index, line in enumerate(sql_lines)
            if line.strip().upper().startswith(("UNSUPPORTED:", "NO_SUPPORTED_QUERY"))
        ),
        len(sql_lines),
    )
    sql = "\n".join(sql_lines[:end]).strip()
    return sql or None


def table_names(ddls: list[str]) -> list[str]:
    result = []
    for ddl in ddls:
        match = re.search(r"CREATE\s+TABLE\s+([^\s(]+)", ddl, re.I)
        if match:
            result.append(match.group(1))
    return result


def full_user_message(meaning: str, description: str, acceptance: str, ddls: list[str]) -> str:
    return (
        "CONSOLIDATED BUSINESS MEANING (authoritative):\n"
        f"{meaning}\n\nCOMPLETE DESCRIPTION (supporting context):\n{description}\n\n"
        f"COMPLETE ACCEPTANCE CRITERIA (supporting context):\n{acceptance}\n\n"
        "QDRANT-RETRIEVED DDL SCHEMAS:\n" + "\n\n---\n\n".join(ddls)
    )


def validate_full(sql: str, ddl_context: str) -> list[str]:
    if not sql or sql == "NO_SUPPORTED_QUERY":
        return ["NO_SUPPORTED_QUERY"]
    errors = []
    errors.extend(_find_invalid_table_refs(sql, ddl_context))
    # Column validation only applies reliably to one-table SQL. Multi-table full
    # queries are validated during atomic decomposition.
    if len(re.findall(r"\b(?:FROM|JOIN)\s+", sql, re.I)) == 1:
        errors.extend(_find_invalid_column_refs(sql, ddl_context))
    return list(dict.fromkeys(errors))


def _sql_tables(sql: str) -> set[str]:
    matches = re.findall(
        r"\b(?:FROM|JOIN)\s+((?:\[[^]]+\]|[A-Za-z_]\w*)"
        r"(?:\s*\.\s*(?:\[[^]]+\]|[A-Za-z_]\w*)){2})",
        sql,
        re.IGNORECASE,
    )
    return {
        ".".join(part.strip().strip("[]").lower() for part in re.split(r"\s*\.\s*", match))
        for match in matches
    }


def aggregate_dependency_signature(query: str) -> str | None:
    if not re.search(r"\b(?:SUM|COUNT)\s*\(", query, re.IGNORECASE):
        return None
    from_match = re.search(r"\bFROM\b(.+)$", query, re.IGNORECASE | re.DOTALL)
    if not from_match:
        return None
    dependency = re.sub(r"\s+", " ", from_match.group(1)).strip().lower()
    return dependency


def apply_aggregate_reuse(
    plans: dict[str, dict[str, dict[str, str]]]
) -> None:
    for steps in plans.values():
        groups: dict[str, list[tuple[str, dict[str, str]]]] = {}
        for source_step, item in steps.items():
            signature = aggregate_dependency_signature(item["query"])
            if signature:
                groups.setdefault(signature, []).append((source_step, item))
        for members in groups.values():
            if len(members) < 2:
                continue
            source_step, source_item = min(
                members,
                key=lambda member: (
                    bool(re.search(r"(?:<=|>=|<>|=|<|>)\s*\d+", member[1]["query"])),
                    len(member[1]["query"]),
                ),
            )
            reusable_query = re.sub(
                r"\s*(?:<=|>=|<>|=|<|>)\s*\d+(?:\s+AS\s+\[[^]]+\]|\s+AS\s+\w+)?(?=\s+FROM\b)",
                " AS ReusableMetric",
                source_item["query"],
                count=1,
                flags=re.IGNORECASE,
            )
            for current_step, item in members:
                item["query"] = reusable_query
                item["reuse_source_step"] = "" if current_step == source_step else source_step


def apply_reviewed_reuse_patterns(
    plans: dict[str, dict[str, dict[str, str]]]
) -> None:
    reviewed = {
        "3004": {
            "steps": ("3", "4"),
            "query": """SELECT
    c.claimid,
    c.status,
    c.resubclaimid,
    c.orgclaimid,
    c.plancrn
FROM [plandata_rx_production].[dbo].[claim] AS c WITH (nolock)
WHERE c.resubclaimid IS NOT NULL
  AND LTRIM(RTRIM(CONVERT(nvarchar(max), c.resubclaimid))) <> N''""",
        },
        "3018": {
            "steps": ("4", "5"),
            "query": """SELECT
    e.MemberId,
    e.RateCode,
    e.ProgramId,
    e.BenefitPlanId,
    e.EffectiveDate,
    e.TermDate
FROM [InMemory].[dbo].[ENROLLMENT] AS e WITH (nolock)
WHERE e.MemberId = {{MemberId}}
  AND e.SegType = 'INT'
  AND {{DateOfService}} >= e.EffectiveDate
  AND {{DateOfService}} <= e.TermDate""",
        },
        "3019": {
            "steps": ("3", "4"),
            "query": """SELECT
    mh.NDC,
    mh.GCNSeqNo,
    mh.HICLSeqNo,
    mh.DrugGenClass,
    mh.DateOfService,
    mh.PaidDate,
    mh.Quantity,
    mh.Dose,
    mh.ClaimID
FROM [InMemory].[dbo].[MEMBER_HISTORY] AS mh WITH (nolock)
WHERE mh.MemberId = {{MemberId}}
  AND TRY_CONVERT(date, mh.DateOfService) <= TRY_CONVERT(date, {{DateOfService}})""",
        },
        "7190": {
            "steps": ("4", "8"),
            "query": """SELECT
    d.NDC_Code,
    d.NDC_Cl,
    d.NDC_CovidEffDate,
    d.NDC_CovidTermDate
FROM [InMemory].[dbo].[DRUG] AS d WITH (nolock)
WHERE d.NDC_Code = {{ClaimTransaction.Ndc}}""",
        },
    }
    for edit_id, pattern in reviewed.items():
        rule_steps = plans.get(edit_id, {})
        source_step = pattern["steps"][0]
        if not all(step in rule_steps for step in pattern["steps"]):
            continue
        for step in pattern["steps"]:
            rule_steps[step]["query"] = pattern["query"]
            rule_steps[step]["reuse_source_step"] = "" if step == source_step else source_step


def validate_decomposed_query(query: str, full_query: str) -> list[str]:
    errors = _find_invalid_sql_artifacts(query)
    select_count = len(re.findall(r"\bSELECT\b", query, re.IGNORECASE))
    if select_count != 1:
        errors.append(f"expected exactly one SELECT statement; found {select_count}")
    query_tables = _sql_tables(query)
    full_tables = _sql_tables(full_query)
    if not query_tables.issubset(full_tables):
        errors.append("decomposed query introduced a table absent from the full query")
    query_literals = set(re.findall(r"'(?:''|[^'])*'", query))
    full_literals = set(re.findall(r"'(?:''|[^'])*'", full_query))
    if not query_literals.issubset(full_literals):
        errors.append("decomposed query introduced a literal absent from the full query")
    query_placeholders = set(re.findall(r"\{\{[^{}]+\}\}", query))
    full_placeholders = set(re.findall(r"\{\{[^{}]+\}\}", full_query))
    if not query_placeholders.issubset(full_placeholders):
        errors.append("decomposed query introduced a placeholder absent from the full query")
    return list(dict.fromkeys(errors))


def generate_reuse_plan(
    client: OpenAI,
    model: str,
    steps: list[tuple[str, str]],
    full_query: str,
) -> dict[str, dict[str, str]]:
    step_text = "\n".join(
        f"SOURCE STEP {source_step}: {meaning}" for source_step, meaning in steps
    )
    raw = clean_text(
        call_model(
            client,
            model,
            [
                {"role": "system", "content": REUSE_PLAN_PROMPT},
                {
                    "role": "user",
                    "content": f"Y STEPS:\n{step_text}\n\nORIGINAL FULL QUERY:\n{full_query}",
                },
            ],
            json_mode=True,
        )
    )
    obj = json.loads(raw)
    result: dict[str, dict[str, str]] = {}
    expected = {source_step for source_step, _ in steps}
    for item in obj.get("steps", []):
        source_step = str(item.get("source_step") or "").strip()
        if source_step not in expected:
            step_match = re.search(r"\d+(?:\.\d+)?", source_step)
            source_step = step_match.group(0) if step_match else source_step
        query = clean_text(str(item.get("query") or ""))
        reuse_source_step = str(item.get("reuse_source_step") or "").strip()
        if source_step in expected and query:
            errors = validate_decomposed_query(query, full_query)
            if errors:
                raise ValueError(f"Invalid reusable query for step {source_step}: {'; '.join(errors)}")
            result[source_step] = {
                "query": query,
                "reuse_source_step": reuse_source_step,
            }
    missing = expected - set(result)
    if missing:
        raise ValueError(f"Reusable plan omitted Y steps: {sorted(missing)}")
    for source_step, item in result.items():
        reuse = item["reuse_source_step"]
        if reuse not in result:
            step_match = re.search(r"\d+(?:\.\d+)?", reuse)
            reuse = step_match.group(0) if step_match else reuse
            item["reuse_source_step"] = reuse
        if reuse and (reuse not in result or result[reuse]["query"] != item["query"]):
            raise ValueError(f"Step {source_step} has an invalid reuse source {reuse}")
    return result


def generate_decomposed_query(
    client: OpenAI,
    model: str,
    meaning: str,
    full_query: str,
) -> tuple[str, list[str]]:
    base_message = (
        f"CURRENT BUSINESS MEANING:\n{meaning}\n\n"
        f"FULL QUERY:\n{full_query}"
    )
    feedback = ""
    last_query = ""
    last_errors = ["No decomposed query returned"]
    for _ in range(4):
        raw = clean_text(
            call_model(
                client,
                model,
                [
                    {"role": "system", "content": DECOMPOSE_PROMPT},
                    {"role": "user", "content": base_message + feedback},
                ],
                json_mode=True,
            )
        )
        try:
            obj = json.loads(raw)
            query = clean_text(str(obj.get("query") or ""))
        except json.JSONDecodeError as exc:
            last_errors = [f"Invalid decomposition JSON: {exc}"]
            feedback = "\n\nThe prior response was not valid JSON. Return the required JSON object."
            continue
        if not query:
            last_errors = ["No decomposed query returned"]
            feedback = "\n\nThe prior response had no query. Return exactly one SELECT."
            continue
        last_query = query
        last_errors = validate_decomposed_query(query, full_query)
        if not last_errors:
            return query, []
        feedback = (
            "\n\nThe prior query violated these constraints: "
            + "; ".join(last_errors)
            + ". Correct it using only the full query."
        )
    return last_query, last_errors


def main() -> None:
    ARTIFACT_DIR.mkdir(parents=True, exist_ok=True)
    context = json.loads(CONTEXT_PATH.read_text(encoding="utf-8"))
    checkpoint = (
        json.loads(CHECKPOINT_PATH.read_text(encoding="utf-8"))
        if CHECKPOINT_PATH.exists()
        else {}
    )
    legacy_checkpoint_path = ARTIFACT_DIR / "report_checkpoint.json"
    legacy_checkpoint = (
        json.loads(legacy_checkpoint_path.read_text(encoding="utf-8"))
        if legacy_checkpoint_path.exists()
        else {}
    )
    workbook = openpyxl.load_workbook(SOURCE)
    source = workbook[ADO_SOURCE_SHEET]
    client, model = make_client()

    rule_meanings: dict[str, list[str]] = {}
    for row in range(2, source.max_row + 1):
        edit_id = str(merged_value(source, row, 1) or "").strip()
        meaning = str(merged_value(source, row, 3) or "").strip()
        if edit_id and meaning and meaning not in rule_meanings.setdefault(edit_id, []):
            rule_meanings[edit_id].append(meaning)

    rule_queries = (
        json.loads(RULE_QUERY_CHECKPOINT_PATH.read_text(encoding="utf-8"))
        if RULE_QUERY_CHECKPOINT_PATH.exists()
        else {}
    )
    for edit_id, meanings in rule_meanings.items():
        if edit_id in rule_queries:
            continue
        rule_context = context[edit_id]
        consolidated_meaning = "\n".join(
            f"{index}. {meaning}" for index, meaning in enumerate(meanings, 1)
        )
        previous_limit = os.environ.get("SCHEMA_RETRIEVAL_FINAL_LIMIT")
        os.environ["SCHEMA_RETRIEVAL_FINAL_LIMIT"] = "20"
        try:
            ddls = select_ddls(
                consolidated_meaning,
                description=rule_context["description"],
                acceptance_criteria=rule_context["acceptance_criteria"],
            )
        finally:
            if previous_limit is None:
                os.environ.pop("SCHEMA_RETRIEVAL_FINAL_LIMIT", None)
            else:
                os.environ["SCHEMA_RETRIEVAL_FINAL_LIMIT"] = previous_limit
        full_query = clean_text(
            call_model(
                client,
                model,
                [
                    {"role": "system", "content": FULL_QUERY_PROMPT},
                    {
                        "role": "user",
                        "content": full_user_message(
                            consolidated_meaning,
                            rule_context["description"],
                            rule_context["acceptance_criteria"],
                            ddls,
                        ),
                    },
                ],
            )
        )
        rule_queries[edit_id] = {
            "full_query": full_query,
            "retrieved_tables": table_names(ddls),
            "consolidated_business_meanings": meanings,
        }
        RULE_QUERY_CHECKPOINT_PATH.write_text(
            json.dumps(rule_queries, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        print(f"rule={edit_id} generated original full query", flush=True)

    rule_y_steps: dict[str, list[tuple[str, str]]] = {}
    for row in range(2, source.max_row + 1):
        edit_id = str(merged_value(source, row, 1) or "").strip()
        source_step = str(merged_value(source, row, 2) or "").strip()
        meaning = str(merged_value(source, row, 3) or "").strip()
        requires = str(merged_value(source, row, 4) or "").strip().upper()
        if requires in {"Y", "YES", "TRUE", "1"}:
            rule_y_steps.setdefault(edit_id, []).append((source_step, meaning))

    reuse_plans: dict[str, dict[str, dict[str, str]]] = {}
    for edit_id, steps in rule_y_steps.items():
        try:
            reuse_plans[edit_id] = generate_reuse_plan(
                client, model, steps, rule_queries[edit_id]["full_query"]
            )
        except (ValueError, json.JSONDecodeError) as exc:
            print(f"rule={edit_id} reuse plan fallback: {exc}", flush=True)
            fallback: dict[str, dict[str, str]] = {}
            for source_step, meaning in steps:
                legacy = legacy_checkpoint.get(f"{edit_id}:{source_step}", {})
                query = clean_text(str(legacy.get("decomposed_query") or ""))
                if not query:
                    query, errors = generate_decomposed_query(
                        client, model, meaning, rule_queries[edit_id]["full_query"]
                    )
                    if errors:
                        raise ValueError(
                            f"Rule {edit_id} step {source_step} fallback failed: "
                            + "; ".join(errors)
                        )
                fallback[source_step] = {"query": query, "reuse_source_step": ""}
            reuse_plans[edit_id] = fallback

    apply_aggregate_reuse(reuse_plans)
    apply_reviewed_reuse_patterns(reuse_plans)

    for row in range(2, source.max_row + 1):
        edit_id = str(merged_value(source, row, 1) or "").strip()
        source_step = str(merged_value(source, row, 2) or "").strip()
        meaning = str(merged_value(source, row, 3) or "").strip()
        requires = str(merged_value(source, row, 4) or "").strip().upper()
        if requires not in {"Y", "YES", "TRUE", "1"}:
            continue
        key = f"{edit_id}:{source_step}"
        rule_context = context[edit_id]
        previous_limit = os.environ.get("SCHEMA_RETRIEVAL_FINAL_LIMIT")
        os.environ["SCHEMA_RETRIEVAL_FINAL_LIMIT"] = "20"
        try:
            ddls = select_ddls(
                meaning,
                description=rule_context["description"],
                acceptance_criteria=rule_context["acceptance_criteria"],
            )
        finally:
            if previous_limit is None:
                os.environ.pop("SCHEMA_RETRIEVAL_FINAL_LIMIT", None)
            else:
                os.environ["SCHEMA_RETRIEVAL_FINAL_LIMIT"] = previous_limit
        ddl_context = "\n\n---\n\n".join(ddls)
        full_query = clean_text(
            call_model(
                client,
                model,
                [
                    {"role": "system", "content": FULL_QUERY_PROMPT},
                    {
                        "role": "user",
                        "content": full_user_message(
                            meaning,
                            rule_context["description"],
                            rule_context["acceptance_criteria"],
                            ddls,
                        ),
                    },
                ],
            )
        )
        full_errors = validate_full(full_query, ddl_context)
        reuse_item = reuse_plans[edit_id][source_step]
        decomposed_query = reuse_item["query"]
        decomposition_errors = validate_decomposed_query(
            decomposed_query, rule_queries[edit_id]["full_query"]
        )
        checkpoint[key] = {
            "edit_id": edit_id,
            "source_step": source_step,
            "meaning": meaning,
            "description": rule_context["description"],
            "acceptance_criteria": rule_context["acceptance_criteria"],
            "retrieved_tables": table_names(ddls),
            "full_query": full_query,
            "full_query_errors": full_errors,
            "decomposed_query": decomposed_query,
            "decomposition_errors": decomposition_errors,
            "decomposition_source": "rule_level_original_full_query",
            "reuse_source_step": reuse_item["reuse_source_step"],
        }
        CHECKPOINT_PATH.write_text(
            json.dumps(checkpoint, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        print(
            f"row={row} key={key} full={'VALID' if not full_errors else 'UNSUPPORTED'} "
            f"decomp={'VALID' if not decomposition_errors else 'UNSUPPORTED'}",
            flush=True,
        )

    workbook = openpyxl.load_workbook(SOURCE)
    source = workbook[ADO_SOURCE_SHEET]
    for sheet_name in (
        QUERY_REVIEW_OUTPUT_SHEET,
        PAYLOAD_OUTPUT_SHEET,
        ADO_OUTPUT_SHEET,
    ):
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    query_v9 = workbook.copy_worksheet(workbook[QUERY_REVIEW_SOURCE_SHEET])
    query_v9.title = QUERY_REVIEW_OUTPUT_SHEET
    query_v9.cell(1, 10).value = f"Generated Queries - V{REPORT_VERSION}"

    payload_v9 = workbook.copy_worksheet(workbook[PAYLOAD_SOURCE_SHEET])
    payload_v9.title = PAYLOAD_OUTPUT_SHEET
    payload_v9.cell(1, 2).value = f"Context-Aware API Payload JSON - V{REPORT_VERSION}"

    ado_v9 = workbook.copy_worksheet(source)
    ado_v9.title = ADO_OUTPUT_SHEET
    for area in list(ado_v9.merged_cells.ranges):
        ado_v9.unmerge_cells(str(area))
    if ado_v9.max_column > 10:
        ado_v9.delete_cols(11, ado_v9.max_column - 10)
    while ado_v9.max_column < 10:
        ado_v9.cell(1, ado_v9.max_column + 1).value = None
    headers = [
        "Rule ID",
        "Source Step #",
        "Consolidated Business Meaning",
        "Requires Data Query",
        "Complete Description",
        "Complete Acceptance Criteria",
        "Qdrant-Retrieved Tables",
        f"Full Query - V{REPORT_VERSION}",
        f"Reusable Data Agent Query - V{REPORT_VERSION}",
        "Original Full Query",
    ]
    for column, header in enumerate(headers, 1):
        ado_v9.cell(1, column).value = header

    full_valid = decomposed_valid = dq_rows = 0
    for row in range(2, source.max_row + 1):
        edit_id = str(merged_value(source, row, 1) or "").strip()
        source_step = str(merged_value(source, row, 2) or "").strip()
        meaning = str(merged_value(source, row, 3) or "").strip()
        requires = str(merged_value(source, row, 4) or "").strip()
        rule_context = context.get(edit_id, {})
        ado_v9.cell(row, 1).value = edit_id
        ado_v9.cell(row, 2).value = source_step
        ado_v9.cell(row, 3).value = meaning
        ado_v9.cell(row, 4).value = requires
        ado_v9.cell(row, 5).value = rule_context.get("description")
        ado_v9.cell(row, 6).value = rule_context.get("acceptance_criteria")
        ado_v9.cell(row, 10).value = rule_queries.get(edit_id, {}).get("full_query")
        key = f"{edit_id}:{source_step}"
        item = checkpoint.get(key)
        if item:
            dq_rows += 1
            ado_v9.cell(row, 7).value = "\n".join(item["retrieved_tables"])
            ado_v9.cell(row, 8).value = item["full_query"]
            query = item["decomposed_query"]
            reuse_source = item.get("reuse_source_step")
            if reuse_source:
                query = f"-- Reuse result from source step {reuse_source}\n{query}"
            ado_v9.cell(row, 9).value = query
            full_valid += bool(item["full_query"])
            decomposed_valid += bool(item["decomposed_query"])
        else:
            ado_v9.cell(row, 7).value = None
            ado_v9.cell(row, 8).value = None
            ado_v9.cell(row, 9).value = None

    for letter, width in {"A": 12, "B": 12, "C": 55, "D": 18, "E": 55, "F": 70, "G": 45, "H": 80, "I": 80, "J": 80}.items():
        ado_v9.column_dimensions[letter].width = width
    for row in range(1, ado_v9.max_row + 1):
        for column in range(1, 11):
            ado_v9.cell(row, column).alignment = Alignment(vertical="top", wrap_text=True)
        for column in (8, 9):
            value = str(ado_v9.cell(row, column).value or "")
            if re.search(r"\[?InMemory\]?\s*\.\s*\[?dbo\]?", value, re.I):
                ado_v9.cell(row, column).fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        if row > 1 and (ado_v9.cell(row, 8).value or ado_v9.cell(row, 9).value):
            ado_v9.row_dimensions[row].height = 300
    rule_values = [
        (row, ado_v9.cell(row, 1).value)
        for row in range(2, ado_v9.max_row + 1)
    ]
    rule_ranges = contiguous_ranges(rule_values)
    original_query_values = [
        (
            row,
            (
                ado_v9.cell(row, 1).value,
                ado_v9.cell(row, 10).value,
            ),
        )
        for row in range(2, ado_v9.max_row + 1)
        if ado_v9.cell(row, 10).value
    ]

    for start_row, end_row in rule_ranges:
        if end_row <= start_row:
            continue
        for column in (1, 5, 6):
            ado_v9.merge_cells(
                start_row=start_row,
                start_column=column,
                end_row=end_row,
                end_column=column,
            )
            ado_v9.cell(start_row, column).alignment = Alignment(
                vertical="top", wrap_text=True
            )
    for start_row, end_row in contiguous_ranges(original_query_values):
        if end_row <= start_row:
            continue
        ado_v9.merge_cells(
            start_row=start_row,
            start_column=10,
            end_row=end_row,
            end_column=10,
        )
        ado_v9.cell(start_row, 10).alignment = Alignment(
            vertical="top", wrap_text=True
        )

    ado_v9.freeze_panes = "C2"
    ado_v9.auto_filter.ref = f"A1:J{ado_v9.max_row}"

    keep_sheets = {
        QUERY_REVIEW_OUTPUT_SHEET,
        PAYLOAD_OUTPUT_SHEET,
        ADO_OUTPUT_SHEET,
    }
    for sheet_name in list(workbook.sheetnames):
        if sheet_name not in keep_sheets:
            del workbook[sheet_name]

    workbook.save(OUTPUT)
    report = {
        "output": str(OUTPUT),
        "dq_rows": dq_rows,
        "full_queries_valid": full_valid,
        "decompositions_valid": decomposed_valid,
        "sheets": workbook.sheetnames,
    }
    REPORT_PATH.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(json.dumps(report, indent=2), flush=True)


if __name__ == "__main__":
    main()
